import time
from urllib import response
import openpyxl
import vk_api
import requests
import csv
import re

# Имя;Фамилмя;Пол;Город;Год Рождения
# -------------------------------------------------
book_tess = openpyxl.open("Сообщества.xlsx")
sheets = book_tess.sheetnames
data_excel = book_tess[sheets[0]] #Тут можно выбрать лист
# print(data_excel[2][0].value) #Тут можно выбрать столбец и строку

# print(sheet_1[2][0].value)
# print(sheet_1.max_row)
# for row in range(2,sheet_1.max_row):
#     print(sheet_1[row][0].value)
# adress = 'vk.com/'
# goko = sheet_1[3][0].value
# print(goko)


# -------------------------------------------------

# WORK VK

# session = vk_api.VkApi(token="vk1.a.ay4hxPxMh-vHKRe_6TOBFo3kU-d4ywrIsFyJl_Z3C0kPJfg-OlUuNt8a6VFEMX-npPY6tFjS-aM0hRcBrAre4barY2zh5PwhZTCz6y-3-7jo9UVMCeMCob6-OIJR3m2HNK6zDK6W3wetE3u890KrKFdbmFwaDAqRLRjJjpIdlTzfpZVAG7asZSPk9etiWNnm")
session = vk_api.VkApi(token="vk1.a.qS-6u3AvUyMKH_q5038CbWuRHKf_mxa2Cpz9tvWq2OgECseZBq0XW92hL27VkwHYdPf2vAFZAHl5PsycOdfY8TLVuHB6XwSEVThx8oUUQE8woksiiU0qk7cQmQxdGk2XB483-bnv2h8TbDlHuBzDBSuI4a5rKEVjeHCuJWWQ_hkERk1cKUZtop6wsljoP23kqFs0lINfRz1BekLO_U87eQ")
# session = vk_api.VkApi(token="vk1.a.D9hMDG3bCkp5lP6d28EGpcWuXqUOX12Y7DzWfRGkVyi_IyUrrpBLZaUghuiyBJcmjOlH6H9zn7Kc59DFIm5ASzYfoM6x2ygHORDJCaOF3soOU2A3h6xTyDdo2Nl--X1Dqn-XOkNBp1Lh9IlyW41RwPcfvDTusTKJuckr_KHCR7Bvbij0-mbI3OYDK70juI93Z0mIYYD0nXg7zhiizCUtaw")
vk = session.get_api()
# def get_user_status(group_id):
#     status = session.method("status.get", {"group_id": group_id})
#     print(status["text"])
# get_user_status(goko)
# status = session.method("status.get", {"group_id": goko})
# f = session.method("users.get", {"user_ids":"id175704881"})



def pars_memb(group_id):
    offset = 0
    count = 1000
    has_more_mambers = True
    resoult = []
    while has_more_mambers == True:
        response = session.method("groups.getMembers", {
            "group_id" : group_id,
            "fields" : "sex , city , bdate , domain",
            "offset" : offset,
            "count" : count
            })
        print(offset,len(response["items"]))
        has_more_mambers = bool(len(response["items"]))
        offset += count
        resoult = resoult + response["items"]
        # time.sleep(1)
    
    return resoult

def get_user_city(user):
    try:
        return user["city"]["title"]
    except:
        return None

def get_user_by(user):
    try:
        date = user["bdate"].split(".")
        if len(date) == 3:
            return date[-1]
    except:
        return None


def save_group_members(users):
    with open('Товары для детей.csv','a',encoding='utf8', newline='') as File:  
        writers = csv.writer(File,delimiter=';')
        # writers.writerow(["Имя", "Фамилмя", "Пол", "Город", "Год Рождения"])
        for user in users:
            writers.writerow([
                user.get("first_name",""),
                user.get("last_name",""),
                user.get("domain",""),
                user.get("sex",""),
                get_user_city(user),
                get_user_by(user),
                goko
            ])


for urlrs in range(2,data_excel.max_row):
    goko = data_excel[urlrs][3].value #выбор столбца
    print(goko)
    if re.findall(r"club\d+", goko):
        nums = re.findall(r'\d+', goko)
        nums = [int(i) for i in nums]
        non_club = " ".join(map(str,nums))
        print(non_club)
        members = pars_memb(non_club)
        save_group_members(members)
    elif re.findall(r"public\d+", goko):
        nums = re.findall(r'\d+', goko)
        nums = [int(i) for i in nums]
        non_club = " ".join(map(str,nums))
        print(non_club)
        members = pars_memb(non_club)
        save_group_members(members)
    else:
        try:
            members = pars_memb(goko)
            save_group_members(members)
        except:
            None
# for i in f:
#     print("%s %s Пол: %s Город: %s" % (i["first_name"],i["last_name"],i["sex"],get_user_city(i)))
















# s = session.method("groups.getMembers", {"group_id": "rostovnadonu", "fields": "sex , city"})
# print(s)

# t={} #создаем словарь для хранения данных, получаемых от API VK

# #Далее следует обращение к API с нашими параметрами:
# t = vk.groups.search(q = "androidblog", count = 1000)
# for h in (t): #Сохраняем результаты поиска в файл"users.txt"
#     with open('users.txt','a') as f1:
#         f1.write((str(data['id'][j]) + ';' #ID исходный
#             + str(t[j]['count']) + ';' #Количество найденных пользователей
#             + str(h['id']) + ';' #ID пользователя VK
#             + h['last_name'] + ';' #Фамилия
#             + h['first_name'] + ';' #Имя
#             + h.get('bdate','') + ';' #Дата рождения
#             + h.get('city',{}).get('title','') #У города несколько параметров - нам нужно название: title
#             + ';\n').encode('cp1251', 'replace').decode('cp1251'))#Для удаления нестандартных символов, которые могут вызывать ошибки

