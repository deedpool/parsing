import time
from urllib import response
import openpyxl
import vk_api
import requests
import csv
import re
from openpyxl import Workbook

book_tess = openpyxl.open("Копия Семичаснов паблики.xlsx")
sheets = book_tess.sheetnames
data_excel = book_tess[sheets[0]] #Тут можно выбрать лист
# print(data_excel[2][0].value) #Тут можно выбрать столбец и строку
# -------------------------------------------------

# WORK VK
session = vk_api.VkApi(token="vk1.a.ay4hxPxMh-vHKRe_6TOBFo3kU-d4ywrIsFyJl_Z3C0kPJfg-OlUuNt8a6VFEMX-npPY6tFjS-aM0hRcBrAre4barY2zh5PwhZTCz6y-3-7jo9UVMCeMCob6-OIJR3m2HNK6zDK6W3wetE3u890KrKFdbmFwaDAqRLRjJjpIdlTzfpZVAG7asZSPk9etiWNnm")
vk = session.get_api()

def pars_memb(group_id):
    offset = 0
    count = 1000
    has_more_mambers = True
    resoult = []
    while has_more_mambers == True:
        response = session.method("groups.getMembers", {
            "group_id" : group_id,
            "fields" : "sex , city , bdate , domain",
            "offset" : offset,
            "count" : count
            })
        print(offset,len(response["items"]))
        has_more_mambers = bool(len(response["items"]))
        offset += count
        resoult = resoult + response["items"]
        # time.sleep(1)   
    return resoult

def get_user_city(user):
    try:
        return user["city"]["title"]
    except:
        return None

def get_user_by(user):
    try:
        date = user["bdate"].split(".")
        if len(date) == 3:
            return date[-1]
    except:
        return None


def save_group_members(users):
    with open('z1.csv','a',encoding='utf8', newline='') as File:  
        writers = csv.writer(File,delimiter=';')
        # writers.writerow(["Имя", "Фамилмя", "Пол", "Город", "Год Рождения"])
        for user in users:
            writers.writerow([
                user.get("first_name",""),
                user.get("last_name",""),
                user.get("domain",""),
                user.get("sex",""),
                get_user_city(user),
                get_user_by(user),
                goko
            ])



def data_save_xlsx(users):
    wb = Workbook()
    ws = wb.active
    gg = []
    for user in users:
        gg.writerow([
            user.get("first_name",""),
            user.get("last_name",""),
            user.get("domain",""),
            user.get("sex",""),
            get_user_city(user),
            get_user_by(user),
            goko
        ])
        ws.append(gg)
        wb.save("zzzzzzzz.xlsx")
# stolb_2 = sheet_1[2][1].value
# print(stolb_2)
# re_URL = stolb_2.split("m/")[1]
# s = session.method("groups.getById", {"group_id": re_URL, "fields": "sex , city"})
# print(s)

for urlrs in range(2,data_excel.max_row):
    stolb_2 = data_excel[urlrs][1].value
    goko = stolb_2.split("m/")[1]
    print(goko)

    if re.findall(r"club\d+", goko):
        nums = re.findall(r'\d+', goko)
        nums = [int(i) for i in nums]
        non_club = " ".join(map(str,nums))
        print(non_club)
        try:
            members = pars_memb(non_club)
            save_group_members(members)
            data_save_xlsx(members)
        except vk_api.exceptions.ApiError:
            None
    elif re.findall(r"public\d+", goko):
        nums = re.findall(r'\d+', goko)
        nums = [int(i) for i in nums]
        non_club = " ".join(map(str,nums))
        print(non_club)
        try:
            members = pars_memb(non_club)
            save_group_members(members)
            data_save_xlsx(members)
        except vk_api.exceptions.ApiError:
            None
    else:
        try:
            members = pars_memb(goko)
            data_save_xlsx(members)
            save_group_members(members)
            print(3)
        except vk_api.exceptions.ApiError:
            None        
        except:
            None














# s = session.method("groups.getMembers", {"group_id": "rostovnadonu", "fields": "sex , city"})
# print(s)

# t={} #создаем словарь для хранения данных, получаемых от API VK

# #Далее следует обращение к API с нашими параметрами:
# t = vk.groups.search(q = "androidblog", count = 1000)
# for h in (t): #Сохраняем результаты поиска в файл"users.txt"
#     with open('users.txt','a') as f1:
#         f1.write((str(data['id'][j]) + ';' #ID исходный
#             + str(t[j]['count']) + ';' #Количество найденных пользователей
#             + str(h['id']) + ';' #ID пользователя VK
#             + h['last_name'] + ';' #Фамилия
#             + h['first_name'] + ';' #Имя
#             + h.get('bdate','') + ';' #Дата рождения
#             + h.get('city',{}).get('title','') #У города несколько параметров - нам нужно название: title
#             + ';\n').encode('cp1251', 'replace').decode('cp1251'))#Для удаления нестандартных символов, которые могут вызывать ошибки

